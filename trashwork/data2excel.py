import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

# 创建一个新的Workbook对象或加载一个已存在的Workbook对象
# 例如，创建一个新的Workbook对象
wb = Workbook()
ws = wb.active  # 获取当前活跃的worksheet

# 或者加载一个已存在的Excel文件
# wb = openpyxl.load_workbook('example.xlsx')
# ws = wb.active

# 指定要添加超链接的单元格和URL
cell_coordinate = 'A1'  # 例如，A1单元格
url = 'https://www.example.com'  # 超链接的URL

# 获取单元格对象
cell = ws[cell_coordinate]

# 添加超链接
cell.value = '点击这里'  # 显示在单元格中的文本
cell.hyperlink = url  # 设置超链接的URL
cell.style = "Hyperlink"  # 设置单元格样式为超链接样式（可选）

# 保存工作簿（如果之前是新建的）
wb.save('/Users/zhouzhou/mycode/my-code/trashwork/example_with_hyperlink.xlsx')