import time
import os

import argparse
import sys
import pandas as pd

from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook, Workbook
from lxml import etree


def parse_args():
    p = argparse.ArgumentParser(description='Selenium Excel -> Excel exporter')
    p.add_argument('--input', default='/Users/zhouzhou/mycode/my-code/trashwork/test.xlsx',help='input xlsx file')
    p.add_argument('--sheet', '-s', default='Sheet1', help='sheet name (default first)')
    p.add_argument('--output', '-o', help='output xlsx path (default results.xlsx)', default='results.xlsx')
    p.add_argument('--headless', action='store_true')
    p.add_argument('--dry-run', action='store_true')
    p.add_argument('--timeout', type=int, default=15, help='page load timeout seconds')
    return p.parse_args()

def read_rows(path: str, sheet: str = None):
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    return rows

def write_results_to_xlsx(results, max_xpaths, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'results'
    # header
    headers = ['row', 'project', 'keyword', 'url']
    for i in range(1, max_xpaths+1):
        headers += [f'xpath_{i}', f'xpath_{i}_extractor', f'xpath_{i}_value']
    ws.append(headers)
    for r in results:
        row = [r.get('row'), r.get('project'), r.get('keyword'), r.get('url')]
        for xp in r.get('xpath_values', []):
            row += [xp.get('xpath'), xp.get('extractor'), xp.get('value')]
        # pad
        while len(row) < len(headers):
            row.append('')
        ws.append(row)
    wb.save(out_path)
    print(f'Wrote results to {out_path}')



def main():
    
    args = parse_args()
    rows = read_rows(args.input, args.sheet)
    if not rows:
        print('No rows read', file=sys.stderr)
        sys.exit(2)
    for row in rows:
        print(row)

    Edge_op = Options()
    # 屏蔽被控制
    Edge_op.add_experimental_option('useAutomationExtension', False)
    Edge_op.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
    driver = webdriver.Edge(options=Edge_op)

    #操作网站
    driver.get("https://www.bilibili.com/")
    time.sleep(5)  #等待5秒钟
    all_windows = driver.window_handles
    search_input = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/div[1]/div/div/form/div[1]/input')
    search_input.send_keys(rows[0][0])
    search_button = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/div[1]/div/div/form/div[2]')
    search_button.click()
    time.sleep(5)  #等待5秒钟
    new_window_handle = None
    for handle in driver.window_handles:
        if handle not in all_windows:  # 新窗口的句柄在当前句柄列表之外
            new_window_handle = handle
            break
    # 切换到新窗口
    driver.switch_to.window(new_window_handle)
    # write_results_to_xlsx(results, max_xpaths, args.output)

    parent_div = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[2]/div/div/div/div[3]/div')
    parent_div_text = parent_div.text.split('\n')


    parent_div_text2d = [parent_div_text[i:i+6] for i in range(0, len(parent_div_text), 6)]
    # 将列表转换为DataFrame
    df = pd.DataFrame(parent_div_text2d, columns=['播放量', '弹幕数', '视频时长', '标题', 'UP主', '发布时间'])

    # 将DataFrame写入Excel文件
    df.to_excel('output.xlsx', index=False)
    # for i in range(0, len(parent_div_text),6):
        
        # print(f'播放量: {parent_div_text[i]}, 视频时长: {parent_div_text[i+2]}, 标题: {parent_div_text[i+3]}, UP主: {parent_div_text[i+4]}, 发布时间: {parent_div_text[i+5]}')



if __name__ == '__main__':
    main()


